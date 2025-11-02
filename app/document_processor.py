import os
import tempfile
from typing import List, Dict, Any
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_core.documents import Document
from langchain_cohere import CohereEmbeddings
from langchain_pinecone import Pinecone
from pinecone import Pinecone as PineconeClient, ServerlessSpec
from dotenv import load_dotenv
import hashlib
import json
import pandas as pd
import base64
from PIL import Image

load_dotenv()

class DocumentProcessor:
    def __init__(self):
        # Initialize Pinecone client
        self.pinecone_api_key = os.environ.get("PINECONE_API_KEY")
        self.pinecone_environment = os.environ.get("PINECONE_ENVIRONMENT", "us-east-1")
        self.pinecone_index_name = os.environ.get("PINECONE_INDEX_NAME", "index-azure-openai-3072")
        
        # Validate required environment variables
        if not self.pinecone_api_key:
            error_msg = """
            ❌ PINECONE_API_KEY environment variable is required!
            
            To fix this issue:
            
            1. In Azure Portal:
               - Go to your Web App
               - Click on "Configuration" 
               - Under "Application settings", click "New application setting"
               - Name: PINECONE_API_KEY
               - Value: your_pinecone_api_key_here
               - Click OK, then Save
            
            2. Or locally, add to your .env file:
               PINECONE_API_KEY=your_pinecone_api_key_here
            
            3. Get your API key from: https://app.pinecone.io/
            """
            raise ValueError(error_msg)
        
        self.pinecone_client = PineconeClient(
            api_key=self.pinecone_api_key,
            environment=self.pinecone_environment
        )
        
        # Initialize embeddings using Azure OpenAI
        try:
            from langchain_openai import AzureOpenAIEmbeddings
            
            # Get Azure OpenAI configuration with defaults
            azure_deployment = os.environ.get("AZ_OPENAI_EMBEDDING_DEPLOYMENT", "text-embedding-3-large")
            azure_endpoint = os.environ.get("AZ_OPENAI_ENDPOINT")
            azure_api_key = os.environ.get("AZ_OPENAI_API_KEY")
            azure_api_version = os.environ.get("AZ_OPENAI_API_VERSION", "2024-12-01-preview")
            
            # Validate required Azure OpenAI variables
            if not azure_endpoint or not azure_api_key:
                error_msg = """
                ❌ Azure OpenAI credentials are missing!
                
                To fix this issue:
                
                1. In Azure Portal:
                   - Go to your Web App
                   - Click on "Configuration"
                   - Under "Application settings", add these settings:
                   - Name: AZ_OPENAI_ENDPOINT
                     Value: your_azure_openai_endpoint_here
                   - Name: AZ_OPENAI_API_KEY
                     Value: your_azure_openai_api_key_here
                   - Click OK, then Save
                
                2. Or locally, add to your .env file:
                   AZ_OPENAI_ENDPOINT=your_azure_openai_endpoint_here
                   AZ_OPENAI_API_KEY=your_azure_openai_api_key_here
                
                3. Get your Azure OpenAI credentials from:
                   - Azure Portal -> Azure OpenAI -> Your resource -> Keys and Endpoint
                """
                print("Azure OpenAI credentials not found, falling back to Cohere embeddings")
                raise ValueError(error_msg)
            
            self.embeddings = AzureOpenAIEmbeddings(
                azure_deployment=azure_deployment,
                azure_endpoint=azure_endpoint,
                api_key=azure_api_key,
                api_version=azure_api_version
            )
            print(f"Using Azure OpenAI embeddings: {azure_deployment}")
        except Exception as e:
            print(f"Error initializing Azure OpenAI embeddings: {e}")
            # Fallback to Cohere as backup
            try:
                self.embeddings = CohereEmbeddings(model="embed-multilingual-v3.0")
                print("Fallback: Using Cohere embed-multilingual-v3.0 model (1024 dimensions)")
            except Exception as e2:
                print(f"Error initializing fallback embeddings: {e2}")
                self.embeddings = CohereEmbeddings(model="multilingual-22-12")
                print("WARNING: Using Cohere multilingual-22-12 model (768 dimensions) - this may not match index dimensions")
        
        # Initialize text splitter
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=1000,
            chunk_overlap=200,
            length_function=len,
        )
        
        # Initialize vector store
        self.vectorstore = None
        self._initialize_vectorstore()
    
    def _initialize_vectorstore(self):
        """Initialize the Pinecone vector store"""
        try:
            # Check if index exists
            if self.pinecone_index_name not in self.pinecone_client.list_indexes().names():
                print(f"Index '{self.pinecone_index_name}' not found. Creating new index...")
                # Create index with correct dimensions for Azure OpenAI text-embedding-3-large (3072 dimensions)
                self.pinecone_client.create_index(
                    name=self.pinecone_index_name,
                    dimension=3072,
                    metric="cosine",
                    spec=ServerlessSpec(
                        cloud="aws",
                        region="us-east-1"
                    )
                )
                print(f"Index '{self.pinecone_index_name}' created successfully")
            
            # Initialize vector store with existing index using the new API
            self.vectorstore = Pinecone.from_existing_index(
                index_name=self.pinecone_index_name,
                embedding=self.embeddings
            )
            print(f"Vector store initialized successfully with index '{self.pinecone_index_name}'")
            
        except Exception as e:
            print(f"Error initializing vector store: {e}")
            # Try alternative initialization for newer Pinecone API
            try:
                # Try using the newer Pinecone API structure
                from pinecone import Pinecone as NewPineconeClient
                new_client = NewPineconeClient(api_key=self.pinecone_api_key)
                
                # Get the index
                index = new_client.Index(self.pinecone_index_name)
                
                # Initialize vector store with the new client
                self.vectorstore = Pinecone(
                    index=index,
                    embedding_function=self.embeddings,
                    text_key="text"
                )
                print(f"Vector store initialized successfully with new API for index '{self.pinecone_index_name}'")
                
            except Exception as e2:
                print(f"Failed to initialize with new API as well: {e2}")
                raise Exception(f"Failed to initialize Pinecone vector store: {str(e)}")
    
    def _generate_document_id(self, content: str) -> str:
        """Generate a unique ID for the document based on its content"""
        return hashlib.md5(content.encode()).hexdigest()
    
    def _flatten_json(self, obj, parent_key='', sep='_'):
        """Flatten nested JSON object into key-value pairs as text"""
        items = []
        
        if isinstance(obj, dict):
            for key, value in obj.items():
                # Skip ID fields and coordinate-like fields
                if any(skip_word in key.lower() for skip_word in ['id', '_id', 'coord', 'coordinate', 'x', 'y', 'z', 'lat', 'lon', 'latitude', 'longitude']):
                    continue
                new_key = f"{parent_key}{sep}{key}" if parent_key else key
                items.extend(self._flatten_json(value, new_key, sep))
        elif isinstance(obj, list):
            for i, value in enumerate(obj):
                new_key = f"{parent_key}{sep}{i}" if parent_key else str(i)
                items.extend(self._flatten_json(value, new_key, sep))
        else:
            # Convert primitive values to text
            if obj is not None and obj != '':
                items.append(f"{parent_key}: {obj}")
        
        return items
    
    def _process_json_file(self, file_path: str, metadata: Dict[str, Any]) -> str:
        """Process JSON file and extract all key-value pairs as flattened text"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            
            # Flatten JSON structure and extract text
            flattened_items = self._flatten_json(json_data)
            content = '\n'.join(flattened_items)
            
            print(f"JSON processed: {len(flattened_items)} key-value pairs extracted")
            return content
            
        except Exception as e:
            raise Exception(f"Error processing JSON file: {e}")
    
    def _process_xlsx_file(self, file_path: str, metadata: Dict[str, Any]) -> str:
        """Process XLSX file and extract headers and row values"""
        try:
            try:
                import openpyxl
            except ImportError as e:
                raise Exception(f"openpyxl library not available. Please ensure it's installed: {e}")
            
            try:
                workbook = openpyxl.load_workbook(file_path)
            except Exception as e:
                raise Exception(f"Failed to load Excel file. Ensure it's a valid .xlsx file: {e}")
            all_text = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows_data = []
                
                # Get headers (first row)
                headers = []
                for cell in sheet[1]:  # Assuming first row contains headers
                    if cell.value is not None:
                        headers.append(str(cell.value))
                
                if headers:
                    rows_data.append(f"Sheet '{sheet_name}' headers: {', '.join(headers)}")
                
                # Process data rows
                for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
                    row_text = []
                    for i, cell_value in enumerate(row):
                        if cell_value is not None:
                            header = headers[i] if i < len(headers) else f"Column_{i+1}"
                            row_text.append(f"{header}: {cell_value}")
                    
                    if row_text:
                        rows_data.append(' | '.join(row_text))
                
                if rows_data:
                    all_text.extend(rows_data)
            
            content = '\n'.join(all_text)
            print(f"XLSX processed: {len(workbook.sheetnames)} sheets, extracted data")
            return content
            
        except Exception as e:
            raise Exception(f"Error processing XLSX file: {e}")
    
    def _process_image_file(self, file_path: str, metadata: Dict[str, Any]) -> str:
        """Process image file using basic text extraction to avoid network issues"""
        try:
            from PIL import Image
            import io
            
            # Open and process image
            with Image.open(file_path) as img:
                file_extension = os.path.splitext(file_path)[1].lower()
                image_data = img.info
                
                # Add image metadata
                metadata.update({
                    "width": img.width,
                    "height": img.height,
                    "format": img.format,
                    "size_bytes": os.path.getsize(file_path)
                })
            
            # Extract basic information from image without API calls for now
            # This avoids network issues with GPT-4o vision API
            image_info = f"""
            Image Analysis: {os.path.basename(file_path)}
            Format: {file_extension}
            Dimensions: {metadata['width']}x{metadata['height']} pixels
            Size: {metadata['size_bytes']} bytes
            
            Content Description:
            - This is a {file_extension} image file
            - File type: {file_extension}
            - Image dimensions: {metadata['width']}x{metadata['height']}
            - File size: {metadata['size_bytes']} bytes
            - Upload timestamp: Current processing time
            
            Note: Image content analysis requires manual review of the visual elements.
            For OCR text extraction, consider using dedicated OCR services.
            """
            
            print(f"Image processed: {file_extension}, {metadata['width']}x{metadata['height']}, basic analysis complete")
            return image_info.strip()
                
        except Exception as e:
            raise Exception(f"Error processing image file: {e}")
    
    def _process_csv_file(self, file_path: str, metadata: Dict[str, Any]) -> str:
        """Process CSV file and include headers with row data"""
        try:
            # Check if pandas is available
            try:
                import pandas as pd
            except ImportError as e:
                raise Exception(f"pandas library not available. Please ensure it's installed: {e}")
            
            # Try different encodings and delimiters
            encodings = ['utf-8', 'latin-1', 'cp1252']
            delimiters = [',', ';', '\t']
            
            df = None
            for encoding in encodings:
                try:
                    for delimiter in delimiters:
                        df = pd.read_csv(file_path, encoding=encoding, delimiter=delimiter)
                        break
                    if df is not None:
                        break
                except:
                    continue
            
            if df is None:
                df = pd.read_csv(file_path)  # Last resort with default settings
            
            # Convert DataFrame to text with headers
            rows_text = []
            
            # Add headers information
            headers_text = f"Headers: {', '.join(df.columns.tolist())}"
            rows_text.append(headers_text)
            
            # Process each row
            for idx, row in df.iterrows():
                row_data = []
                for col in df.columns:
                    if pd.notna(row[col]):
                        row_data.append(f"{col}: {row[col]}")
                
                if row_data:
                    row_text = f"Row {idx + 1}: {' | '.join(row_data)}"
                    rows_text.append(row_text)
            
            content = '\n'.join(rows_text)
            print(f"CSV processed: {len(df)} rows, {len(df.columns)} columns")
            return content
            
        except Exception as e:
            raise Exception(f"Error processing CSV file: {e}")
    
    def process_text(self, text: str, metadata: Dict[str, Any] = None) -> List[Document]:
        """Process raw text into chunks"""
        if metadata is None:
            metadata = {}
        
        # Generate document ID
        doc_id = self._generate_document_id(text)
        metadata["doc_id"] = doc_id
        metadata["source_type"] = "text"
        
        # Split text into chunks
        chunks = self.text_splitter.split_text(text)
        
        # Create documents
        documents = []
        for i, chunk in enumerate(chunks):
            chunk_metadata = metadata.copy()
            chunk_metadata["chunk_id"] = f"{doc_id}_{i}"
            documents.append(Document(page_content=chunk, metadata=chunk_metadata))
        
        return documents
    
    def process_file(self, file_path: str, metadata: Dict[str, Any] = None) -> List[Document]:
        """Process a file based on its extension"""
        if metadata is None:
            metadata = {}
        
        file_extension = os.path.splitext(file_path)[1].lower()
        
        if file_extension == '.txt':
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            metadata["source_type"] = "text_file"
            metadata["file_name"] = os.path.basename(file_path)
            return self.process_text(content, metadata)
        
        elif file_extension == '.pdf':
            try:
                from pypdf import PdfReader
                reader = PdfReader(file_path)
                content = ""
                for page in reader.pages:
                    content += page.extract_text() + "\n"
                metadata["source_type"] = "pdf_file"
                metadata["file_name"] = os.path.basename(file_path)
                metadata["num_pages"] = len(reader.pages)
                return self.process_text(content, metadata)
            except Exception as e:
                raise Exception(f"Error processing PDF file: {e}")
        
        elif file_extension == '.json':
            content = self._process_json_file(file_path, metadata)
            metadata["source_type"] = "json_file"
            metadata["file_name"] = os.path.basename(file_path)
            return self.process_text(content, metadata)
        
        elif file_extension in ['.xlsx', '.xls']:
            content = self._process_xlsx_file(file_path, metadata)
            metadata["source_type"] = "excel_file"
            metadata["file_name"] = os.path.basename(file_path)
            return self.process_text(content, metadata)
        
        elif file_extension == '.csv':
            content = self._process_csv_file(file_path, metadata)
            metadata["source_type"] = "csv_file"
            metadata["file_name"] = os.path.basename(file_path)
            return self.process_text(content, metadata)
        
        elif file_extension in ['.jpg', '.jpeg', '.png']:
            content = self._process_image_file(file_path, metadata)
            metadata["source_type"] = "image_file"
            metadata["file_name"] = os.path.basename(file_path)
            return self.process_text(content, metadata)
        
        else:
            raise Exception(f"Unsupported file type: {file_extension}")
    
    def add_documents_to_vectorstore(self, documents: List[Document]):
        """Add processed documents to the vector store"""
        try:
            # Add documents to Pinecone
            self.vectorstore.add_documents(documents)
            return True
        except Exception as e:
            print(f"Error adding documents to vector store: {e}")
            return False
    
    def search_documents(self, query: str, k: int = 5):
        """Search for relevant documents"""
        if self.vectorstore is None:
            return []
        
        try:
            results = self.vectorstore.similarity_search(query, k=k)
            return results
        except Exception as e:
            print(f"Error searching documents: {e}")
            return []
    
    def get_retriever(self):
        """Get the retriever for the RAG chain"""
        if self.vectorstore is None:
            raise Exception("Vector store not initialized")
        
        # Use synchronous retriever to avoid session closure issues
        return self.vectorstore.as_retriever(search_kwargs={"k": 5})
    
    def clear_documents(self, doc_id: str = None):
        """Clear documents from the vector store"""
        # This is a simplified implementation
        # In a production environment, you would want to implement proper document deletion
        print(f"Clear documents functionality not fully implemented. Doc ID: {doc_id}")
        return True
